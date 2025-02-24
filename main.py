import os 
import tkinter as tk 
import threading 
import time 
from tkinter import ttk, messagebox 
from tkinter import filedialog 
from PIL import Image, ImageTk 
from openpyxl import load_workbook

# Global variables
current_row = None
current_image = None
retry_count = 0
FOLDER_PATH = ""
EXCEL_PATH = ""
image_size = (1536, 864)
navigation_history = []  # New history tracking list

# Function to prompt the user to select window size
def choose_window_size():
    size_prompt = tk.Tk()
    size_prompt.title("Choose Window Size")
    size_prompt.geometry("300x150")
    
    def set_normal_size():
        global image_size, window_size
        window_size = "1550x1200"
        image_size = (1536, 864)
        size_prompt.destroy()

    def set_small_size():
        global image_size, window_size
        window_size = "1163x900"
        image_size = (1152, 648)
        size_prompt.destroy()

    tk.Label(size_prompt, text="Select Window Size").pack(pady=10)
    
    normal_button = tk.Button(size_prompt, text="Normal (1550x1200)", command=set_normal_size)
    normal_button.pack(pady=5)
    
    small_button = tk.Button(size_prompt, text="Small (1163x900)", command=set_small_size)
    small_button.pack(pady=5)
    
    size_prompt.grab_set()
    size_prompt.mainloop()

choose_window_size()

window = tk.Tk()
window.title("Zooniverse Validator v.1.08")
window.geometry(window_size)

def choose_folder_and_file():
    global FOLDER_PATH, EXCEL_PATH, wb, ws

    FOLDER_PATH = filedialog.askdirectory(title="Select Image Folder")
    if not FOLDER_PATH:
        messagebox.showerror("Error", "You must select an image folder.")
        return False

    EXCEL_PATH = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if not EXCEL_PATH:
        messagebox.showerror("Error", "You must select an Excel file.")
        return False

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open Excel file: {e}")
        return False

    return True

def find_next_unvalidated():
    for row in range(2, ws.max_row + 1):
        status = ws[f'A{row}'].value
        if status is None:
            status = "FALSE"
        status = str(status).strip().upper()
        
        if status == "FALSE":
            return row
    return None

original_info_label = tk.Label(window, text="", justify=tk.LEFT, wraplength=1000)
original_info_label.grid(row=5, column=0, columnspan=4)

def find_original_classification(filename):
    original_classifications = []
    for row in range(2, ws.max_row + 1):
        original_filename = ws[f'L{row}'].value
        if original_filename == filename:
            species1_orig = ws[f'M{row}'].value or "NONE"
            species1_count_orig = ws[f'N{row}'].value or "NONE"
            species2_orig = ws[f'O{row}'].value or "NONE"
            species2_count_orig = ws[f'P{row}'].value or "NONE"
            original_classifications.append(
                (species1_orig, species1_count_orig, species2_orig, species2_count_orig)
            )
    return original_classifications

def display_original_classifications(classifications):
    original_info_label.config(text="")
    column_widths = [30, 20, 30, 20]
    total_width = sum(column_widths) + len(column_widths) - 1
    
    header = f"{'Species 1':<{column_widths[0]}} {'Count 1':<{column_widths[1]}} {'Species 2':<{column_widths[2]}} {'Count 2':<{column_widths[3]}}\n"
    info_text = header
    
    for (species1, count1, species2, count2) in classifications:
        line = f"{species1:<{column_widths[0]}} {str(count1):<{column_widths[1]}} {species2:<{column_widths[2]}} {str(count2):<{column_widths[3]}}\n"
        info_text += line
    
    original_info_label.config(text=info_text, font=("Courier", 11))

def show_image(image_path):
    global current_image
    try:
        image = Image.open(image_path)
        image = image.resize(image_size, Image.Resampling.LANCZOS)
        current_image = ImageTk.PhotoImage(image)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open image: {e}")
    image_label.config(image=current_image)

def find_image_in_subfolders(filename):
    start_time = time.time()
    target_filename = filename.lower()
    
    for root, dirs, files in os.walk(FOLDER_PATH):
        for file in files:
            if file.lower() == target_filename:
                return os.path.join(root, file)
    return None

def get_total_rows_in_column_c():
    total_rows = 0
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value is not None:
            total_rows += 1
    return total_rows

def get_true_rows_in_column_A():
    true_rows = 0
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value == "TRUE":
            true_rows += 1
    return true_rows

image_info_label = tk.Label(window, text="", font=("Courier", 10))
image_info_label.grid(row=2, column=0, columnspan=4)

def update_image_info(filename, row_number):
    total_rows = get_total_rows_in_column_c()
    true_rows = get_true_rows_in_column_A()
    image_info_label.config(text=f"{filename} ({true_rows}/{total_rows})")

def load_next_image():
    global current_row, retry_count
    
    if current_row is None:
        messagebox.showinfo("Done", "All images validated!")
        return

    original_info_label.config(text="Loading image...")
    window.update_idletasks()

    def load_image_thread():
        global retry_count
        
        print(f"Starting to load image for row {current_row}...")
        filename = ws[f'C{current_row}'].value
        if filename:
            image_path = find_image_in_subfolders(filename)
            if image_path:
                window.after(1, lambda: show_image(image_path))
                retry_count = 0
                window.after(1, lambda: update_image_info(filename, current_row - 1))
                original_classifications = find_original_classification(filename)
                if original_classifications:
                    window.after(1, lambda: display_original_classifications(original_classifications))
                else:
                    window.after(1, lambda: original_info_label.config(text="No original classifications found for this image."))
            else:
                retry_count += 1
                if retry_count >= 3:
                    window.after(1, lambda: messagebox.showinfo("Stopping", "Image could not be found - Please verify filename in excel sheet and image folder."))
                    window.after(1, window.quit)
                    return
                window.after(1, load_next_image)
        else:
            retry_count += 1
            if retry_count >= 3:
                window.after(1, lambda: messagebox.showinfo("Stopping", "Image could not be found - Please verify filename in excel sheet and image folder."))
                window.after(1, window.quit)
                return
            window.after(1, load_next_image)
    
    threading.Thread(target=load_image_thread).start()

def go_back():
    global current_row, navigation_history
    
    if not navigation_history:
        messagebox.showinfo("Info", "Cannot go back further.")
        return

    # Get last validated row from history
    previous_row = navigation_history.pop()
    
    # Reset its validation status
    ws[f'A{previous_row}'].value = "FALSE"
    wb.save(EXCEL_PATH)
    
    # Set current row and load
    current_row = previous_row
    load_image(current_row)

def load_image(row):
    if row is None:
        messagebox.showinfo("Done", "All images validated!")
        return

    original_info_label.config(text="Loading image...")
    window.update_idletasks()

    filename = ws[f'C{row}'].value
    if filename:
        image_path = find_image_in_subfolders(filename)
        if image_path:
            show_image(image_path)
            update_image_info(filename, row - 1)
            original_classifications = find_original_classification(filename)
            if original_classifications:
                display_original_classifications(original_classifications)
            else:
                original_info_label.config(text="No original classifications found for this image.")
        else:
            messagebox.showerror("Error", f"File {filename} not found!")
    else:
        messagebox.showerror("Error", "No filename found in the current row.")

back_button = tk.Button(window, text="Back", command=go_back)
back_button.grid(row=4, column=1, padx=0, sticky=tk.N)

def save_and_next():
    global current_row, navigation_history
    if current_row is None:
        return
    
    # Save current data
    ws[f'D{current_row}'].value = species1_var.get()
    ws[f'F{current_row}'].value = species2_var.get()
    ws[f'E{current_row}'].value = species1_count_var.get()
    ws[f'G{current_row}'].value = species2_count_var.get()
    
    # Mark current row as validated
    ws[f'A{current_row}'].value = "TRUE"
    wb.save(EXCEL_PATH)
    
    # Add to navigation history before finding next
    navigation_history.append(current_row)  # Track validated row
    
    # Find next unvalidated row
    current_row = find_next_unvalidated()
    
    if current_row is None:
        messagebox.showinfo("Done", "All images validated!")
    else:
        load_next_image()



class AutocompleteCombobox(ttk.Combobox):
    def set_completion_list(self, completion_list):
        self._completion_list = sorted(completion_list)
        self._hits = []
        self._hit_index = 0
        self.position = 0
        self.bind('<KeyRelease>', self.handle_keyrelease)
        self['values'] = self._completion_list

    def autocomplete(self, delta=0):
        if delta:
            self._hit_index += delta
            if self._hit_index == len(self._hits):
                self._hit_index = 0
            elif self._hit_index == -1:
                self._hit_index = len(self._hits) - 1
        if self._hits:
            self.delete(0, tk.END)
            self.insert(0, self._hits[self._hit_index])
            self.select_range(self.position, tk.END)

    def handle_keyrelease(self, event):
        if event.keysym in ('BackSpace', 'Left', 'Right', 'Up', 'Down'):
            return

        self.position = self.index(tk.END)
        _input = self.get()

        if _input == '':
            self._hits = self._completion_list
        else:
            self._hits = [item for item in self._completion_list if item.lower().startswith(_input.lower())]

        if self._hits:
            self._hit_index = 0
            self.autocomplete()

window.grid_columnconfigure(0, weight=0)
window.grid_columnconfigure(1, weight=0)
window.grid_columnconfigure(2, weight=0)
window.grid_columnconfigure(3, weight=0)

species1_label = tk.Label(window, text="Species 1: ")
species1_label.grid(row=1, column=0, sticky="e", padx=0, pady=5)

species1_count_label = tk.Label(window, text="Species 1 Count: ")
species1_count_label.grid(row=1, column=2, sticky="e", padx=0, pady=5)

species2_label = tk.Label(window, text="Species 2: ")
species2_label.grid(row=2, column=0, sticky="e", padx=0, pady=5)

species2_count_label = tk.Label(window, text="Species 2 Count: ")
species2_count_label.grid(row=2, column=2, sticky="e", padx=0, pady=5)

species_options = ["NONE", "HUMAN", "MULEDEER", "DESERTCOTTONTAIL", "COYOTE", "RODENT", "VIRGINIAOPOSSUM", "EASTERNFOXSQUIRREL", "CALIFORNIAGROUNDSQUIRREL", "STRIPEDSKUNK", "BIRD", "BLACKBEAR", "RACCOON", "BOBCAT", "SNAKE", "DOMESTICDOG", "DOMESTICCAT"]

count_options = ["NONE", "1", "2", "3", "4ORMORE"]

species1_var = tk.StringVar(value="NONE")
species2_var = tk.StringVar(value="NONE")

species1_dropdown = AutocompleteCombobox(window, textvariable=species1_var, width=30)
species1_dropdown.set_completion_list(species_options)
species1_dropdown.grid(row=1, column=1, padx=0, pady=5, sticky="w")

species2_dropdown = AutocompleteCombobox(window, textvariable=species2_var, width=30)
species2_dropdown.set_completion_list(species_options)
species2_dropdown.grid(row=2, column=1, padx=0, pady=5, sticky="w")

species1_count_var = tk.StringVar(value="NONE")
species1_count_dropdown = ttk.Combobox(window, textvariable=species1_count_var, values=count_options, width=20, state="normal")
species1_count_dropdown.grid(row=1, column=3, padx=0, pady=5, sticky="w")

species2_count_var = tk.StringVar(value="NONE")
species2_count_dropdown = ttk.Combobox(window, textvariable=species2_count_var, values=count_options, width=20, state="normal")
species2_count_dropdown.grid(row=2, column=3, padx=0, pady=5, sticky="w")

image_label = tk.Label(window)
image_label.grid(row=0, column=0, columnspan=4)

next_button = tk.Button(window, text="Next", command=save_and_next)
next_button.grid(row=4, column=1, padx=0, columnspan=2)

set_all_none_button = tk.Button(window, text="Set All to None", command=lambda: (
    species1_var.set("NONE"),
    species2_var.set("NONE"),
    species1_count_var.set("NONE"),
    species2_count_var.set("NONE")
))
set_all_none_button.grid(row=3, column=1, columnspan=2)

if choose_folder_and_file():
    current_row = find_next_unvalidated()
    if current_row:
        load_next_image()
    else:
        messagebox.showinfo("Done", "All images are already validated.")
else:
    window.destroy()

window.mainloop()